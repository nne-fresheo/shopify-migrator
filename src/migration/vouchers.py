from __future__ import annotations

import csv
import logging
import secrets
import string
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import click

from .client import ShopifyClient

logger = logging.getLogger(__name__)

# Alphabet for the random part of a discount code. Excludes ambiguous characters
# (0/O, 1/I) so codes stay easy to read and type.
_CODE_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
_CODE_SUFFIX_LEN = 8

# CSV report columns, in order.
_REPORT_FIELDS = ["email", "status", "code", "customer_id", "discount_id", "error"]

# Columns for the update report.
_UPDATE_REPORT_FIELDS = ["email", "code", "status", "discount_id", "error"]

_DISCOUNT_CREATE_MUTATION = """
mutation CreateVoucher($basicCodeDiscount: DiscountCodeBasicInput!) {
  discountCodeBasicCreate(basicCodeDiscount: $basicCodeDiscount) {
    codeDiscountNode { id }
    userErrors { field message }
  }
}
"""

_DISCOUNT_FIND_QUERY = """
query FindVoucher($query: String!) {
  codeDiscountNodes(first: 50, query: $query) {
    edges {
      node {
        id
        codeDiscount {
          __typename
          ... on DiscountCodeBasic {
            title
            codes(first: 1) { edges { node { code } } }
          }
        }
      }
    }
  }
}
"""

_DISCOUNT_BY_CODE_QUERY = """
query VoucherByCode($code: String!) {
  codeDiscountNodeByCode(code: $code) { id }
}
"""

_DISCOUNT_UPDATE_MUTATION = """
mutation UpdateVoucher($id: ID!, $basicCodeDiscount: DiscountCodeBasicInput!) {
  discountCodeBasicUpdate(id: $id, basicCodeDiscount: $basicCodeDiscount) {
    codeDiscountNode { id }
    userErrors { field message }
  }
}
"""


def read_emails(path: Path, column: str = "email") -> list[str]:
    """Read email addresses from the given column of an .xlsx file.

    Locates the header cell matching ``column`` (case-insensitive) on the first
    worksheet, then collects the non-empty values below it, normalized to
    lowercase and de-duplicated while preserving order.
    """
    # Imported lazily so the rest of the CLI works without openpyxl installed.
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise click.ClickException(f"{path}: the spreadsheet is empty.")

        col_idx = _find_column(header, column)
        if col_idx is None:
            found = ", ".join(str(h) for h in header if h is not None) or "(none)"
            raise click.ClickException(
                f"{path}: no '{column}' column found. Header columns: {found}"
            )

        seen: set[str] = set()
        emails: list[str] = []
        for row in rows:
            if col_idx >= len(row):
                continue
            value = row[col_idx]
            if value is None:
                continue
            email = str(value).strip().lower()
            if not email or email in seen:
                continue
            seen.add(email)
            emails.append(email)
        return emails
    finally:
        wb.close()


def _find_column(header: tuple, column: str) -> Optional[int]:
    target = column.strip().lower()
    for idx, cell in enumerate(header):
        if cell is not None and str(cell).strip().lower() == target:
            return idx
    return None


def generate_code(prefix: str) -> str:
    """Build a unique, human-readable discount code: ``<PREFIX>-<RANDOM>``."""
    suffix = "".join(secrets.choice(_CODE_ALPHABET) for _ in range(_CODE_SUFFIX_LEN))
    return f"{prefix}-{suffix}"


class VoucherGenerator:
    """Creates a per-email, customer-restricted fixed-amount voucher in Shopify.

    For each email it ensures a Customer record exists (creating one if needed),
    then creates a basic code discount (GraphQL ``discountCodeBasicCreate``)
    restricted to that customer with a unique code. The discount can target
    one-time purchases, subscriptions, or both; for subscriptions
    ``recurring_cycles`` controls how many billing cycles it applies to
    (1 = first order only, 0 = every recurring order). Re-runs are idempotent:
    a discount whose title already matches an email is reused, not recreated.
    """

    def __init__(
        self,
        dest_client: ShopifyClient,
        amount: float,
        days: int,
        prefix: str = "FRESHEO",
        applies_one_time: bool = True,
        applies_subscription: bool = True,
        recurring_cycles: int = 1,
        dry_run: bool = False,
    ) -> None:
        self.dest = dest_client
        self.amount = amount
        self.days = days
        self.prefix = prefix
        self.applies_one_time = applies_one_time
        self.applies_subscription = applies_subscription
        self.recurring_cycles = recurring_cycles
        self.dry_run = dry_run

    def _title(self, email: str) -> str:
        return f"{self.prefix} – {email}"

    async def find_customer(self, email: str) -> Optional[dict]:
        """Return the Shopify customer exactly matching ``email``, or None."""
        response = await self.dest.get(
            "customers/search.json", params={"query": f"email:{email}"}
        )
        for customer in response.get("customers", []):
            if str(customer.get("email", "")).strip().lower() == email:
                return customer
        return None

    async def ensure_customer(self, email: str) -> dict:
        """Find the customer by email, creating a minimal record if missing."""
        existing = await self.find_customer(email)
        if existing:
            return existing
        if self.dry_run:
            logger.info(f"[DRY RUN] would create customer for {email}")
            return {"id": None, "email": email}
        response = await self.dest.post("customers.json", {"customer": {"email": email}})
        return response["customer"]

    async def find_existing_discount(self, email: str) -> Optional[dict]:
        """Return a basic code discount already created for this email, or None.

        Searches by title and exact-matches in Python, since discount search is
        tokenized/fuzzy. Returns ``{"id": <gid>, "code": <code>}`` on a hit.
        """
        title = self._title(email)
        data = await self.dest.graphql(
            _DISCOUNT_FIND_QUERY, variables={"query": f'title:"{title}"'}
        )
        for edge in data.get("codeDiscountNodes", {}).get("edges", []):
            node = edge.get("node", {})
            discount = node.get("codeDiscount") or {}
            if discount.get("title") != title:
                continue
            code_edges = discount.get("codes", {}).get("edges", [])
            code = code_edges[0]["node"]["code"] if code_edges else ""
            return {"id": node.get("id", ""), "code": code}
        return None

    def _discount_input(self, email: str, customer_id) -> dict:
        now = datetime.now(timezone.utc)
        customer_gets: dict = {
            "value": {
                "discountAmount": {
                    "amount": f"{self.amount:.2f}",
                    "appliesOnEachItem": False,
                }
            },
            "items": {"all": True},
            "appliesOnOneTimePurchase": self.applies_one_time,
            "appliesOnSubscription": self.applies_subscription,
        }
        payload: dict = {
            "title": self._title(email),
            "startsAt": now.isoformat(),
            "endsAt": (now + timedelta(days=self.days)).isoformat(),
            "appliesOncePerCustomer": True,
            "customerSelection": {
                "customers": {"add": [f"gid://shopify/Customer/{customer_id}"]}
            },
            "customerGets": customer_gets,
        }
        # recurringCycleLimit is only meaningful when the discount hits subscriptions.
        if self.applies_subscription:
            payload["recurringCycleLimit"] = self.recurring_cycles
        return payload

    async def create_voucher(self, email: str) -> dict:
        """Create (or reuse) a voucher for one email. Returns a report row."""
        row: dict = {
            "email": email,
            "status": "",
            "code": "",
            "customer_id": "",
            "discount_id": "",
            "error": "",
        }
        try:
            # Idempotency: reuse a discount already issued for this email.
            existing = await self.find_existing_discount(email)
            if existing:
                row.update(
                    status="exists",
                    discount_id=existing["id"],
                    code=existing.get("code", ""),
                )
                logger.info(f"[vouchers] {email}: already has a voucher, skipping")
                return row

            customer = await self.ensure_customer(email)
            row["customer_id"] = "" if customer.get("id") is None else str(customer["id"])

            code = generate_code(self.prefix)
            row["code"] = code

            if self.dry_run:
                row["status"] = "would-create"
                logger.info(
                    f"[DRY RUN] would create voucher {code} (-{self.amount:.2f} EUR, "
                    f"{self.days}d) for {email}"
                )
                return row

            payload = self._discount_input(email, customer["id"])
            payload["code"] = code
            data = await self.dest.graphql(
                _DISCOUNT_CREATE_MUTATION, variables={"basicCodeDiscount": payload}
            )
            result = data.get("discountCodeBasicCreate", {})
            errors = result.get("userErrors") or []
            if errors:
                raise RuntimeError(f"discountCodeBasicCreate userErrors: {errors}")

            row["discount_id"] = result.get("codeDiscountNode", {}).get("id", "")
            row["status"] = "created"
            logger.info(
                f"[vouchers] {email}: created {code} (-{self.amount:.2f} EUR, "
                f"{self.days}d, discount={row['discount_id']})"
            )
            return row

        except Exception as exc:
            row["status"] = "failed"
            row["error"] = str(exc)
            logger.error(f"[vouchers] {email}: FAILED — {exc}")
            return row

    async def run(self, emails: list[str]) -> list[dict]:
        """Process every email; one failure never aborts the batch."""
        logger.info(f"[vouchers] starting ({len(emails)} emails)")
        rows = [await self.create_voucher(email) for email in emails]
        logger.info("[vouchers] done")
        return rows


def write_report(rows: list[dict], path: Path, fields: Optional[list[str]] = None) -> None:
    """Write result rows to a CSV file using the given column order."""
    fieldnames = fields or _REPORT_FIELDS
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def read_report(path: Path) -> list[dict]:
    """Read a previously written voucher report (CSV) into a list of row dicts."""
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


class VoucherUpdater:
    """Applies a partial update to existing vouchers listed in a report CSV.

    Only the fields supplied to the constructor are changed; everything else on
    each discount is left untouched (``discountCodeBasicUpdate`` is a partial
    update). Each voucher is resolved by its ``code`` via
    ``codeDiscountNodeByCode``, so it works regardless of whether the discount
    was originally created through the REST or GraphQL API.
    """

    def __init__(
        self,
        dest_client: ShopifyClient,
        *,
        amount: Optional[float] = None,
        days: Optional[int] = None,
        purchase_type: Optional[str] = None,
        recurring_cycles: Optional[int] = None,
        usage_limit: Optional[int] = None,
        once_per_customer: Optional[bool] = None,
        dry_run: bool = False,
    ) -> None:
        self.dest = dest_client
        self.amount = amount
        self.days = days
        self.purchase_type = purchase_type
        self.recurring_cycles = recurring_cycles
        self.usage_limit = usage_limit
        self.once_per_customer = once_per_customer
        self.dry_run = dry_run

    def build_input(self) -> dict:
        """Build the partial DiscountCodeBasicInput from the supplied fields."""
        payload: dict = {}
        customer_gets: dict = {}

        if self.amount is not None:
            customer_gets["value"] = {
                "discountAmount": {
                    "amount": f"{self.amount:.2f}",
                    "appliesOnEachItem": False,
                }
            }
        if self.purchase_type is not None:
            customer_gets["appliesOnOneTimePurchase"] = self.purchase_type in (
                "one-time",
                "both",
            )
            customer_gets["appliesOnSubscription"] = self.purchase_type in (
                "subscription",
                "both",
            )
        if customer_gets:
            payload["customerGets"] = customer_gets

        if self.recurring_cycles is not None:
            payload["recurringCycleLimit"] = self.recurring_cycles
        if self.days is not None:
            ends = datetime.now(timezone.utc) + timedelta(days=self.days)
            payload["endsAt"] = ends.isoformat()
        if self.usage_limit is not None:
            payload["usageLimit"] = self.usage_limit
        if self.once_per_customer is not None:
            payload["appliesOncePerCustomer"] = self.once_per_customer

        return payload

    def has_changes(self) -> bool:
        return bool(self.build_input())

    async def update_one(self, row: dict, update_input: dict) -> dict:
        """Resolve one voucher by code and apply the partial update."""
        code = (row.get("code") or "").strip()
        result = {
            "email": row.get("email", ""),
            "code": code,
            "status": "",
            "discount_id": "",
            "error": "",
        }
        if not code:
            result["status"] = "skipped"
            result["error"] = "no code in report row"
            return result
        try:
            data = await self.dest.graphql(_DISCOUNT_BY_CODE_QUERY, variables={"code": code})
            node = data.get("codeDiscountNodeByCode")
            if not node:
                result["status"] = "not-found"
                logger.warning(f"[update] {code}: no discount found for code, skipping")
                return result
            result["discount_id"] = node["id"]

            if self.dry_run:
                result["status"] = "would-update"
                logger.info(f"[DRY RUN] would update {code} ({node['id']})")
                return result

            response = await self.dest.graphql(
                _DISCOUNT_UPDATE_MUTATION,
                variables={"id": node["id"], "basicCodeDiscount": update_input},
            )
            update = response.get("discountCodeBasicUpdate", {})
            errors = update.get("userErrors") or []
            if errors:
                raise RuntimeError(f"discountCodeBasicUpdate userErrors: {errors}")
            result["status"] = "updated"
            logger.info(f"[update] {code}: updated ({node['id']})")
            return result

        except Exception as exc:
            result["status"] = "failed"
            result["error"] = str(exc)
            logger.error(f"[update] {code}: FAILED — {exc}")
            return result

    async def run(self, rows: list[dict]) -> list[dict]:
        """Apply the update to every report row; one failure never aborts the batch."""
        update_input = self.build_input()
        logger.info(f"[update] starting ({len(rows)} vouchers)")
        results = [await self.update_one(row, update_input) for row in rows]
        logger.info("[update] done")
        return results
